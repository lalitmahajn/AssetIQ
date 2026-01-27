from typing import Annotated, Any, Union

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import select

from apps.plant_backend.deps import require_perm
from apps.plant_backend.models import Asset, TimelineEvent
from common_core.db import PlantSessionLocal

router = APIRouter(prefix="/ui/assets", tags=["ui-assets"])


@router.get("/{asset_id}/history")
def get_asset_history(
    asset_id: str,
    limit: int = 10,
    user: Annotated[Any, Depends(require_perm("ticket.view"))] = None,
):
    db = PlantSessionLocal()
    try:
        # Fetch events for this asset (STOP, TICKET, etc.)
        q = (
            select(TimelineEvent)
            .where(TimelineEvent.asset_id == asset_id)
            .order_by(TimelineEvent.occurred_at_utc.desc())
            .limit(limit)
        )

        events = db.execute(q).scalars().all()

        return [
            {
                "id": e.id,
                "type": e.event_type,
                "occurred_at": e.occurred_at_utc.isoformat(),
                "payload": e.payload_json,
            }
            for e in events
        ]
    finally:
        db.close()


@router.post("/import")
def import_assets(
    file: Annotated[UploadFile, File(...)],
    user: Annotated[Any, Depends(require_perm("asset.edit"))] = None,
):
    import openpyxl
    from datetime import datetime
    from uuid import uuid4

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Invalid file format. Please upload .xlsx")

    db = PlantSessionLocal()
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active

        # Headers: Asset Code, Name, Category, Parent Asset Code, Location Area, Criticality
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"imported": 0, "failed": 0, "errors": ["Empty file"]}

        headers = [str(h).lower() for h in rows[0] if h]
        required = ["asset code", "name", "category"]
        for r in required:
            if r not in headers:
                raise HTTPException(400, f"Missing required header: {r}")

        # Map headers to indices
        idx = {h: i for i, h in enumerate(headers)}

        imported = 0
        updated = 0
        failed = 0
        errors = []

        # Pass 1: Upsert Assets
        # Store parent mappings for Pass 2: {child_code: parent_code}
        parent_map = {}

        for i, row in enumerate(rows[1:], start=2):
            try:
                # Safe access
                def get_val(h):
                    return (
                        str(row[idx[h]]).strip() if h in idx and row[idx[h]] is not None else None
                    )

                code = get_val("asset code")
                name = get_val("name")
                category = get_val("category")

                if not code or not name or not category:
                    failed += 1
                    errors.append(f"Row {i}: Missing required fields")
                    continue

                # Check exist
                asset = db.query(Asset).filter_by(asset_code=code).first()
                if not asset:
                    # Create
                    asset = Asset(
                        id=str(uuid4()),
                        site_code="P01",  # Default to P01 or user's site?
                        asset_code=code,
                        name=name,
                        category=category,
                        created_at_utc=datetime.utcnow(),
                        status="active",
                        is_active=True,
                    )
                    db.add(asset)
                    imported += 1
                else:
                    # Update
                    asset.name = name
                    asset.category = category
                    updated += 1

                # Update optional fields
                loc = get_val("location area")
                if loc:
                    asset.location_area = loc

                crit = get_val("criticality")
                if crit:
                    asset.criticality = crit.lower()
                    asset.is_critical = crit.lower() == "high" or crit.lower() == "critical"

                # Store parent for pass 2
                p_code = get_val("parent asset code")
                if p_code:
                    parent_map[code] = p_code

            except Exception as e:
                failed += 1
                errors.append(f"Row {i}: {str(e)}")

        db.flush()

        # Pass 2: Link Parents
        # Fetch all potential parent assets
        parent_codes = list(set(parent_map.values()))
        parent_assets = db.query(Asset).filter(Asset.asset_code.in_(parent_codes)).all()
        code_to_id = {a.asset_code: a.id for a in parent_assets}

        # Now update children
        child_codes = list(parent_map.keys())
        children = db.query(Asset).filter(Asset.asset_code.in_(child_codes)).all()

        for child in children:
            p_code = parent_map.get(child.asset_code)
            if p_code and p_code in code_to_id:
                child.parent_id = code_to_id[p_code]

        db.commit()

        return {
            "imported": imported,
            "updated": updated,
            "failed": failed,
            "errors": errors[:20],  # Limit size
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}") from e
    finally:
        db.close()


@router.get("/import/template")
def get_import_template():
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        ["Asset Code", "Name", "Category", "Parent Asset Code", "Location Area", "Criticality"]
    )
    # Sample
    ws.append(["P01-M-001", "Main Feed Pump", "PUMP", "", "Zone A", "High"])
    ws.append(["P01-M-001-MTR", "Pump Motor", "MOTOR", "P01-M-001", "Zone A", "Medium"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=asset_import_template.xlsx"},
    )
