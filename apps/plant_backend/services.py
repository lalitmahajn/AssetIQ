from __future__ import annotations

import uuid
from datetime import datetime, timedelta
from typing import Any

from sqlalchemy import select

from apps.plant_backend.models import (
    Asset,
    AuditLog,
    EmailQueue,
    EventOutbox,
    MasterItem,
    MasterType,
    ReasonSuggestion,
    ReportRequest,
    StopQueue,
    SystemConfig,
    Ticket,
    TicketActivity,
    TimelineEvent,
    WhatsAppQueue,
)
from common_core.config import settings


def _now() -> datetime:
    return datetime.utcnow()


def _new_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:18]}"


def audit_write(
    db,
    action: str,
    entity_type: str,
    entity_id: str,
    details: dict[str, Any],
    actor_user_id: str | None,
    actor_station_code: str | None,
    request_id: str | None,
) -> None:
    db.add(
        AuditLog(
            site_code=settings.plant_site_code,
            actor_user_id=actor_user_id,
            actor_station_code=actor_station_code,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            request_id=request_id,
            details_json=details,
            created_at_utc=_now(),
        )
    )


def timeline_append(
    db, asset_id: str, event_type: str, payload: dict[str, Any], correlation_id: str
) -> str:
    tid = _new_id("TL")
    db.add(
        TimelineEvent(
            id=tid,
            site_code=settings.plant_site_code,
            asset_id=asset_id,
            event_type=event_type,
            payload_json=payload,
            occurred_at_utc=_now(),
            correlation_id=correlation_id,
            created_at_utc=_now(),
        )
    )
    return tid


def enqueue_email(db, to_email: str, subject: str, body: str) -> None:
    db.add(
        EmailQueue(
            to_email=to_email,
            subject=subject,
            body=body,
            status="PENDING",
            created_at_utc=_now(),
            sent_at_utc=None,
        )
    )


def outbox_add(
    db, entity_type: str, entity_id: str, payload: dict[str, Any], correlation_id: str
) -> None:
    db.add(
        EventOutbox(
            site_code=settings.plant_site_code,
            entity_type=entity_type,
            entity_id=entity_id,
            payload_json=payload,
            correlation_id=correlation_id,
            created_at_utc=_now(),
            sent_at_utc=None,
            retry_count=0,
            next_attempt_at_utc=_now(),
            last_error=None,
        )
    )


def log_ticket_activity(
    db, ticket_id: str, activity_type: str, details: str, actor_id: str | None = None
) -> None:
    db.add(
        TicketActivity(
            ticket_id=ticket_id,
            activity_type=activity_type,
            details=details[:512] if details else None,
            actor_id=actor_id,
            created_at_utc=_now(),
        )
    )


def open_stop(
    db,
    asset_id: str,
    reason: str,
    actor_user_id: str | None,
    actor_station_code: str | None,
    request_id: str | None,
    extra_context: dict = None,
):
    stop_id = _new_id("STOP")
    now = _now()
    db.add(
        StopQueue(
            id=stop_id,
            site_code=settings.plant_site_code,
            asset_id=asset_id,
            reason=reason,
            is_open=True,
            opened_at_utc=now,
            closed_at_utc=None,
            resolution_text=None,
            live_context_json=extra_context,
        )
    )

    corr_stop = f"stop_open:{stop_id}"
    timeline_append(db, asset_id, "STOP_OPEN", {"stop_id": stop_id, "reason": reason}, corr_stop)

    ticket_id = _new_id("TCK")
    sla_due = now + timedelta(minutes=60)
    db.add(
        Ticket(
            id=ticket_id,
            site_code=settings.plant_site_code,
            asset_id=asset_id,
            title=f"Stop: {asset_id} - {reason[:120]}",
            status="OPEN",
            priority="HIGH",
            assigned_to_user_id=None,
            source="AUTO",
            stop_id=stop_id,
            created_at_utc=now,
            sla_due_at_utc=sla_due,
            acknowledged_at_utc=None,
            resolved_at_utc=None,
            resolution_reason=None,
            close_note=None,
        )
    )

    log_ticket_activity(db, ticket_id, "CREATED", f"Auto-generated from Stop {stop_id}", "SYSTEM")

    corr_ticket = f"ticket_open:{ticket_id}"
    timeline_append(
        db, asset_id, "TICKET_OPEN", {"ticket_id": ticket_id, "stop_id": stop_id}, corr_ticket
    )

    enqueue_email(
        db,
        settings.email_maintenance,
        f"[{settings.plant_site_code}] STOP {asset_id} - Ticket {ticket_id}",
        f"Stop opened for asset={asset_id}\nReason={reason}\nTicket={ticket_id}\nSLA Due={sla_due.isoformat()}Z",
    )

    audit_write(
        db,
        "STOP_OPEN",
        "stop_queue",
        stop_id,
        {"asset_id": asset_id, "reason": reason, "ticket_id": ticket_id},
        actor_user_id,
        actor_station_code,
        request_id,
    )

    outbox_add(
        db,
        "timeline_event",
        corr_stop,
        {
            "event_type": "STOP_OPEN",
            "asset_id": asset_id,
            "stop_id": stop_id,
            "reason": reason,
            "occurred_at_utc": now.isoformat() + "Z",
        },
        corr_stop,
    )
    outbox_add(
        db,
        "ticket",
        ticket_id,
        {
            "ticket_id": ticket_id,
            "asset_id": asset_id,
            "stop_id": stop_id,
            "sla_due_at_utc": sla_due.isoformat() + "Z",
            "status": "OPEN",
        },
        corr_ticket,
    )

    return {"stop_id": stop_id, "ticket_id": ticket_id, "sla_due_at_utc": sla_due.isoformat() + "Z"}


def acknowledge_ticket(db, ticket_id: str, actor_user_id: str, request_id: str | None):
    t = db.get(Ticket, ticket_id)
    if not t:
        raise ValueError("TICKET_NOT_FOUND")
    if t.status == "CLOSED":
        raise ValueError("CANNOT_ACK_CLOSED_TICKET")

    if not t.acknowledged_at_utc:
        t.acknowledged_at_utc = _now()
        t.status = "ACK"

        # [NEW] Auto-assign to the person acknowledging if unassigned
        if not t.assigned_to_user_id:
            t.assigned_to_user_id = actor_user_id

        audit_write(db, "TICKET_ACK", "ticket", ticket_id, {}, actor_user_id, None, request_id)
        log_ticket_activity(
            db,
            ticket_id,
            "ACK",
            f"Ticket acknowledged (Assigned to {actor_user_id})",
            actor_user_id,
        )
        timeline_append(
            db,
            t.asset_id,
            "TICKET_ACK",
            {"ticket_id": ticket_id, "assigned_to": actor_user_id},
            f"ticket_ack:{ticket_id}",
        )
        outbox_add(
            db,
            "ticket",
            ticket_id,
            {
                "ticket_id": ticket_id,
                "status": "ACK",
                "assigned_to": actor_user_id,
                "acknowledged_at_utc": t.acknowledged_at_utc.isoformat() + "Z",
            },
            f"ticket_ack:{ticket_id}",
        )
    return t


def resolve_stop(
    db, stop_id: str, resolution_text: str, actor_user_id: str, request_id: str | None
):
    sq = db.get(StopQueue, stop_id)
    if not sq:
        raise ValueError("STOP_NOT_FOUND")
    if not sq.is_open:
        return sq
    sq.is_open = False
    sq.closed_at_utc = _now()
    sq.resolution_text = resolution_text
    audit_write(
        db,
        "STOP_RESOLVE",
        "stop_queue",
        stop_id,
        {"resolution": resolution_text},
        actor_user_id,
        None,
        request_id,
    )
    timeline_append(
        db,
        sq.asset_id,
        "STOP_RESOLVE",
        {"stop_id": stop_id, "resolution": resolution_text},
        f"stop_resolve:{stop_id}",
    )
    outbox_add(
        db,
        "timeline_event",
        f"stop_resolve:{stop_id}",
        {
            "event_type": "STOP_RESOLVE",
            "stop_id": stop_id,
            "asset_id": sq.asset_id,
            "resolution": resolution_text,
            "occurred_at_utc": sq.closed_at_utc.isoformat() + "Z",
            "reason_code": sq.reason,
        },
        f"stop_resolve:{stop_id}",
    )

    # [NEW] Record suggestion if it's not a master reason
    suggestion_record(db, "STOP_REASON", resolution_text, actor_user_id)

    return sq


def create_ticket(
    db,
    title: str,
    asset_id: str,
    priority: str,
    source: str = "MANUAL",
    stop_id: str | None = None,
    actor_id: str | None = None,
    assigned_to: str | None = None,
    dept: str | None = None,
) -> Ticket:
    # [NEW] Smart Asset Creation Logic
    # 1. Case-insensitive lookup
    existing_asset = db.execute(select(Asset).where(Asset.id.ilike(asset_id))).scalars().first()

    final_asset_id = asset_id
    if existing_asset:
        # Use existing casing
        final_asset_id = existing_asset.id
    else:
        # Create new asset
        new_asset = Asset(
            id=asset_id,
            site_code=settings.plant_site_code,
            asset_code=asset_id,  # Default code = ID
            name=asset_id,  # Default name = ID
            category="MACHINE",
            description="Auto-created from manual ticket",
            is_active=True,
            created_at_utc=_now(),
        )
        db.add(new_asset)
        db.flush()  # Ensure ID is available if needed, though we set it explicitly
        final_asset_id = asset_id

    tid = _new_id("TCK")
    now = _now()
    # default SLA 4 hours for manual tickets, 2h for HIGH
    sla_hours = 2 if priority == "HIGH" or priority == "CRITICAL" else 4
    sla_due = now + timedelta(hours=sla_hours)

    t = Ticket(
        id=tid,
        site_code=settings.plant_site_code,
        # Use final_asset_id to ensure FK consistency (even if soft constraint)
        asset_id=final_asset_id,
        title=title,
        status="OPEN",
        priority=priority,
        assigned_to_user_id=assigned_to,
        assigned_dept=dept,
        source=source,
        stop_id=stop_id,
        created_at_utc=now,
        sla_due_at_utc=sla_due,
        acknowledged_at_utc=None,
        resolved_at_utc=None,
        resolution_reason=None,
        close_note=None,
    )
    db.add(t)

    log_ticket_activity(db, tid, "CREATED", f"Ticket created via {source}", actor_id)

    corr = f"ticket_manual:{tid}"
    timeline_append(db, final_asset_id, "TICKET_CREATE", {"ticket_id": tid, "title": title}, corr)
    audit_write(
        db,
        "TICKET_CREATE",
        "ticket",
        tid,
        {
            "asset_id": final_asset_id,
            "title": title,
            "priority": priority,
            "assigned_to": assigned_to,
            "assigned_dept": dept,
        },
        actor_id,
        None,
        None,
    )
    outbox_add(
        db,
        "ticket",
        tid,
        {
            "ticket_id": tid,
            "asset_id": final_asset_id,
            "title": title,
            "status": "OPEN",
            "assigned_to": assigned_to,
        },
        corr,
    )

    # WhatsApp Alert Logic
    try:
        ws_enabled = db.get(SystemConfig, "whatsappEnabled")
        ws_phone = db.get(SystemConfig, "whatsappTargetPhone")
        ws_template = db.get(SystemConfig, "whatsappMessageTemplate")

        if ws_enabled and ws_enabled.config_value is True and ws_phone and ws_phone.config_value:
            # Default template if not set
            raw_msg = "🚀 AssetIQ Ticket Created\nID: {id}\nAsset: {asset_id}\nTitle: {title}\nPriority: {priority}"
            if ws_template and ws_template.config_value:
                raw_msg = str(ws_template.config_value)

            # Format the message
            # Safe substitution to avoid crashes if template has invalid keys
            msg = (
                raw_msg.replace("{id}", str(tid))
                .replace("{asset_id}", str(final_asset_id))
                .replace("{title}", str(title))
                .replace("{priority}", str(priority))
                .replace("{created_at}", str(now.strftime("%Y-%m-%d %H:%M:%S")))
                .replace("{source}", str(source or "Unknown"))
                .replace("{assigned_to}", str(assigned_to or "Unassigned"))
                .replace("{dept}", str(dept or "General"))
                .replace("{sla_due}", str(sla_due.strftime("%Y-%m-%d %H:%M:%S")))
                .replace("{site_code}", str(settings.plant_site_code))
            )

            # Calculate SLA state for conditional routing
            # For newly created tickets, SLA is always OK (just created)
            sla_state = "OK"  # New tickets start with OK state

            db.add(
                WhatsAppQueue(
                    ticket_id=tid,
                    phone_number=str(ws_phone.config_value),
                    message=msg,
                    status="PENDING",
                    sla_state=sla_state,
                    created_at_utc=now,
                )
            )
    except Exception as e:
        # Don't fail ticket creation just because WhatsApp queue failed
        import logging

        logging.getLogger("assetiq").error(f"Failed to queue WhatsApp alert: {e}")

    return t


def close_ticket(
    db,
    ticket_id: str,
    close_note: str,
    resolution_reason: str | None = None,
    actor_id: str | None = None,
) -> Ticket:
    t = db.get(Ticket, ticket_id)
    if not t:
        raise ValueError("TICKET_NOT_FOUND")
    if t.status == "CLOSED":
        return t

    t.status = "CLOSED"
    t.resolved_at_utc = _now()
    t.close_note = close_note
    t.resolution_reason = resolution_reason

    log_ticket_activity(
        db, ticket_id, "CLOSED", f"Closed. Reason: {resolution_reason or 'None'}", actor_id
    )

    audit_write(
        db,
        "TICKET_CLOSE",
        "ticket",
        ticket_id,
        {"close_note": close_note, "reason": resolution_reason},
        actor_id,
        None,
        None,
    )
    timeline_append(
        db,
        t.asset_id,
        "TICKET_CLOSE",
        {"ticket_id": ticket_id, "close_note": close_note},
        f"ticket_close:{ticket_id}",
    )

    # WhatsApp Alert Logic for Closure
    try:
        ws_enabled = db.get(SystemConfig, "whatsappEnabled")
        ws_phone = db.get(SystemConfig, "whatsappTargetPhone")
        ws_template = db.get(SystemConfig, "whatsappCloseMessageTemplate")

        if ws_enabled and ws_enabled.config_value is True and ws_phone and ws_phone.config_value:
            # Default template if not set
            raw_msg = "✅ Ticket Closed\nID: {id}\nNote: {close_note}"
            if ws_template and ws_template.config_value:
                raw_msg = str(ws_template.config_value)

            closed_at_str = (
                t.resolved_at_utc.strftime("%Y-%m-%d %H:%M:%S") if t.resolved_at_utc else "N/A"
            )

            # Format the message
            msg = (
                raw_msg.replace("{id}", str(ticket_id))
                .replace("{asset_id}", str(t.asset_id))
                .replace("{title}", str(t.title))
                .replace("{priority}", str(t.priority))
                .replace("{close_note}", str(close_note))
                .replace("{resolution_reason}", str(resolution_reason or "N/A"))
                .replace("{closed_at}", closed_at_str)
                .replace("{site_code}", str(settings.plant_site_code))
            )

            # Calculate SLA state at closure
            sla_state = "OK"  # Default
            if t.sla_due_at_utc:
                if t.resolved_at_utc and t.resolved_at_utc > t.sla_due_at_utc:
                    sla_state = "BREACHED"
                elif t.resolved_at_utc and t.resolved_at_utc <= t.sla_due_at_utc:
                    sla_state = "OK"

            db.add(
                WhatsAppQueue(
                    ticket_id=ticket_id,
                    phone_number=str(ws_phone.config_value),
                    message=msg,
                    status="PENDING",
                    sla_state=sla_state,
                    created_at_utc=_now(),
                )
            )
    except Exception as e:
        import logging

        logging.getLogger("assetiq").error(f"Failed to queue WhatsApp close alert: {e}")

    outbox_add(
        db,
        "ticket",
        ticket_id,
        {"ticket_id": ticket_id, "status": "CLOSED", "close_note": close_note},
        f"ticket_close:{ticket_id}",
    )
    return t


def assign_ticket(db, ticket_id: str, assigned_user_id: str, actor_id: str | None = None) -> Ticket:
    t = db.get(Ticket, ticket_id)
    if not t:
        raise ValueError("TICKET_NOT_FOUND")

    old = t.assigned_to_user_id
    t.assigned_to_user_id = assigned_user_id

    log_ticket_activity(
        db, ticket_id, "ASSIGNED", f"Assigned to {assigned_user_id} (was {old})", actor_id
    )

    audit_write(
        db,
        "TICKET_ASSIGN",
        "ticket",
        ticket_id,
        {"assigned_to": assigned_user_id},
        actor_id,
        None,
        None,
    )
    timeline_append(
        db,
        t.asset_id,
        "TICKET_ASSIGN",
        {"ticket_id": ticket_id, "assigned_to": assigned_user_id},
        f"ticket_assign:{ticket_id}",
    )
    outbox_add(
        db,
        "ticket",
        ticket_id,
        {"ticket_id": ticket_id, "assigned_to": assigned_user_id},
        f"ticket_assign:{ticket_id}",
    )
    return t


# -----------------------------
# Asset Registry (CMMS Core)
# -----------------------------


def _validate_asset_payload(payload: dict) -> None:
    asset_code = (payload.get("asset_code") or "").strip()
    if not asset_code:
        raise ValueError("asset_code_required")
    if len(asset_code) > 64:
        raise ValueError("asset_code_too_long")
    name = (payload.get("name") or "").strip()
    if not name:
        raise ValueError("name_required")
    category = (payload.get("category") or "").strip()
    if not category:
        raise ValueError("category_required")
    criticality = (payload.get("criticality") or "medium").strip().lower()
    if criticality not in ("low", "medium", "high", "critical"):
        raise ValueError("invalid_criticality")
    status = (payload.get("status") or "active").strip().lower()
    if status not in ("active", "inactive"):
        raise ValueError("invalid_status")
    tags = payload.get("tags") or []
    if not isinstance(tags, list):
        raise ValueError("invalid_tags")


def asset_create(db, payload: dict, actor_user_id: str | None, request_id: str | None):
    _validate_asset_payload(payload)
    site_code = settings.plant_site_code
    asset_code = payload["asset_code"].strip()

    exists = db.execute(
        select(Asset).where(Asset.site_code == site_code, Asset.asset_code == asset_code)
    ).scalar_one_or_none()
    if exists:
        raise ValueError("asset_code_already_exists")

    asset_id = payload.get("id") or _new_id("AST")
    now = _now()

    a = Asset(
        id=asset_id,
        site_code=site_code,
        asset_code=asset_code,
        name=payload["name"].strip(),
        category=payload["category"].strip(),
        parent_id=payload.get("parent_id"),
        criticality=payload.get("criticality", "medium").lower(),  # Deprecated but kept for compat
        is_critical=payload.get("is_critical", False),
        tags=payload.get("tags", []),
        location_area=payload.get("location_area"),
        location_line=payload.get("location_line"),
        status=payload.get("status", "active").lower(),
        created_at_utc=now,
        updated_at_utc=now,
        created_by_user_id=actor_user_id,
    )
    db.add(a)
    db.flush()

    timeline_append(
        db,
        asset_id,
        "ASSET_CREATE",
        {"asset_code": a.asset_code, "name": a.name, "is_critical": a.is_critical},
        f"asset_create:{asset_id}",
    )
    audit_write(
        db,
        "ASSET_CREATE",
        "asset",
        asset_id,
        {"asset_code": asset_code, "is_critical": a.is_critical},
        actor_user_id,
        None,
        request_id,
    )
    outbox_add(
        db,
        "asset",
        asset_id,
        {"asset_id": asset_id, "asset_code": asset_code},
        f"asset_create:{asset_id}",
    )
    return a


def asset_get(db, asset_id: str):
    return db.get(Asset, asset_id)


def asset_tree(db):
    assets = (
        db.execute(
            select(Asset).where(
                Asset.site_code == settings.plant_site_code, Asset.is_active.is_(True)
            )
        )
        .scalars()
        .all()
    )
    by_id = {a.id: a for a in assets}
    children = {}
    for a in assets:
        pid = a.parent_id or "ROOT"
        children.setdefault(pid, []).append(a.id)

    def build(node_id: str):
        if node_id == "ROOT":
            nodes = children.get("ROOT", [])
            return [build(cid) for cid in nodes]
        a = by_id[node_id]
        return {
            "id": a.id,
            "asset_code": a.asset_code,
            "name": a.name,
            "category": a.category,
            "children": [build(cid) for cid in children.get(a.id, [])],
        }

    return {"children": build("ROOT")}


# -------------------------
# Dynamic Masters
# -------------------------


def master_type_list(db, include_inactive: bool = False):
    q = select(MasterType).where(MasterType.site_code == settings.plant_site_code)
    if not include_inactive:
        q = q.where(MasterType.is_active.is_(True))
    return db.execute(q).scalars().all()


def master_item_create(
    db,
    master_type_code: str,
    item_code: str,
    item_name: str,
    meta: dict | None = None,
    actor_id: str | None = None,
):
    now = _now()
    mi = MasterItem(
        site_code=settings.plant_site_code,
        master_type_code=master_type_code,
        item_code=item_code,
        item_name=item_name,
        meta_json=meta or {},
        created_at_utc=now,
        updated_at_utc=now,
    )
    db.add(mi)
    db.flush()
    audit_write(
        db,
        "MASTER_ITEM_CREATE",
        "master_item",
        str(mi.id),
        {"code": item_code},
        actor_id,
        None,
        None,
    )
    return mi


# -------------------------
# Self-learning Suggestions
# -------------------------


def suggestion_record(
    db, master_type_code: str, typed_text: str, actor_id: str | None = None, threshold: int = 5
):
    import re

    norm = re.sub(r"\s+", " ", typed_text.strip().lower())
    if not norm:
        return None

    row = db.execute(
        select(ReasonSuggestion).where(
            ReasonSuggestion.site_code == settings.plant_site_code,
            ReasonSuggestion.master_type_code == master_type_code,
            ReasonSuggestion.normalized_key == norm,
        )
    ).scalar_one_or_none()

    now = _now()
    if not row:
        row = ReasonSuggestion(
            site_code=settings.plant_site_code,
            master_type_code=master_type_code,
            suggested_name=typed_text,
            normalized_key=norm,
            count=1,
            status="pending",
            threshold=threshold,
            created_at_utc=now,
            updated_at_utc=now,
            created_by_user_id=actor_id,
        )
        db.add(row)
    else:
        row.count += 1
        row.updated_at_utc = now
        if row.count >= row.threshold and row.status == "pending":
            row.status = "auto_promoted"
            enqueue_email(
                db,
                settings.email_maintenance,
                f"Suggestion Auto-Promoted: {norm}",
                f"The reason '{typed_text}' has been used {row.count} times.",
            )

    db.flush()
    return row


def suggestion_approve(db, suggestion_id: int, item_code: str, actor_id: str):
    s = db.get(ReasonSuggestion, suggestion_id)
    if not s or s.status in ("approved", "merged"):
        return s

    mi = master_item_create(
        db, s.master_type_code, item_code, s.suggested_name, {"from_suggestion": s.id}, actor_id
    )
    s.status = "approved"
    s.approved_master_item_id = mi.id
    s.reviewed_by_user_id = actor_id
    s.reviewed_at_utc = _now()
    return s


# -----------------------------
# Reporting v1 (manual trigger)
# -----------------------------


def _parse_iso_date(s: str) -> datetime:
    s = (s or "").strip()
    if not s:
        raise ValueError("date is required")
    # Accept YYYY-MM-DD or ISO datetime
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return datetime.fromisoformat(s + "T00:00:00")
    return datetime.fromisoformat(s)


def report_request_create_and_generate_csv(
    db,
    report_type: str,
    date_from: str,
    date_to: str,
    filters: dict,
    actor_user_id: str,
    actor_station_code: str | None,
    request_id: str | None,
) -> ReportRequest:
    import json
    import os

    from apps.plant_backend.models import StopQueue  # Use StopQueue from models.py

    dt_from = _parse_iso_date(date_from)
    dt_to = _parse_iso_date(date_to)
    if dt_to < dt_from:
        raise ValueError("date_to must be >= date_from")

    rr = ReportRequest(
        site_code=settings.plant_site_code,
        report_type=report_type,
        date_from=dt_from,
        date_to=dt_to,
        filters_json=json.dumps(filters or {}, ensure_ascii=False),
        requested_by_user_id=actor_user_id,
        status="requested",
        generated_file_path=None,
        error_message=None,
        created_at_utc=_now(),
        updated_at_utc=_now(),
    )
    db.add(rr)
    db.flush()

    audit_write(
        db,
        "REPORT_REQUEST",
        "report_request",
        str(rr.id),
        {"type": report_type},
        actor_user_id,
        actor_station_code,
        request_id,
    )

    # Generate synchronously
    try:
        vault_root = settings.report_vault_root
        os.makedirs(vault_root, exist_ok=True)
        # safe_type = re.sub(r"[^a-zA-Z0-9_\\-]", "_", report_type)[:64] # Unused

        if report_type == "downtime_by_asset":
            # Generate Excel for downtime_by_asset - shows individual stop records
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Asset_Downtime_{rr.site_code}_{date_str}.xlsx"
            file_path = os.path.join(vault_root, filename)

            stops = (
                db.execute(
                    select(StopQueue)
                    .where(
                        StopQueue.site_code == rr.site_code,
                        StopQueue.opened_at_utc >= dt_from,
                        StopQueue.opened_at_utc <= dt_to,
                    )
                    .order_by(StopQueue.opened_at_utc.desc())
                )
                .scalars()
                .all()
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "Downtime by Asset"

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

            # Headers - individual stop records
            headers = [
                "Site Code",
                "Asset ID",
                "Stop Reason",
                "Started At",
                "Ended At",
                "Duration (Minutes)",
                "Duration (Hours)",
                "Status",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows - individual stops
            row = 2
            for s in stops:
                end_time = s.closed_at_utc
                if end_time:
                    dur_sec = (end_time - s.opened_at_utc).total_seconds()
                    status = "Closed"
                else:
                    dur_sec = (dt_to - s.opened_at_utc).total_seconds()
                    status = "Open"
                    end_time = None  # Will show as blank

                ws.cell(row=row, column=1, value=s.site_code)
                ws.cell(row=row, column=2, value=s.asset_id)
                ws.cell(row=row, column=3, value=s.reason or "N/A")
                ws.cell(row=row, column=4, value=s.opened_at_utc.strftime("%Y-%m-%d %H:%M:%S"))
                ws.cell(
                    row=row,
                    column=5,
                    value=end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time else "Still Open",
                )
                ws.cell(row=row, column=6, value=round(dur_sec / 60, 1))
                ws.cell(row=row, column=7, value=round(dur_sec / 3600, 2))
                ws.cell(row=row, column=8, value=status)
                row += 1

            # Auto-width columns
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 35)

            wb.save(file_path)

        elif report_type == "ticket_performance":
            # Generate PDF for ticket performance metrics
            from collections import Counter, defaultdict

            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Ticket_Performance_{rr.site_code}_{date_str}.pdf"
            file_path = os.path.join(vault_root, filename)

            tickets = (
                db.execute(
                    select(Ticket).where(
                        Ticket.site_code == rr.site_code,
                        Ticket.created_at_utc >= dt_from,
                        Ticket.created_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            total_tickets = len(tickets)
            closed_tickets = [t for t in tickets if t.status == "CLOSED"]
            ack_tickets = [t for t in tickets if t.acknowledged_at_utc]

            # SLA compliance
            sla_met = [
                t
                for t in closed_tickets
                if t.resolved_at_utc and t.sla_due_at_utc and t.resolved_at_utc <= t.sla_due_at_utc
            ]
            sla_compliance = (len(sla_met) / len(closed_tickets) * 100) if closed_tickets else 0

            # Time metrics
            mttr_list = [
                (t.resolved_at_utc - t.created_at_utc).total_seconds() / 3600
                for t in closed_tickets
                if t.resolved_at_utc
            ]
            avg_mttr = sum(mttr_list) / len(mttr_list) if mttr_list else 0

            mtta_list = [
                (t.acknowledged_at_utc - t.created_at_utc).total_seconds() / 60
                for t in ack_tickets
                if t.acknowledged_at_utc
            ]
            avg_mtta = sum(mtta_list) / len(mtta_list) if mtta_list else 0

            # Distributions
            priority_counts = Counter(t.priority for t in tickets)
            status_counts = Counter(t.status for t in tickets)
            dept_counts = Counter(t.assigned_dept or "Unassigned" for t in tickets)
            user_counts = Counter(t.assigned_to_user_id or "Unassigned" for t in tickets)

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "Title",
                parent=styles["Heading1"],
                fontSize=20,
                textColor=colors.HexColor("#1e40af"),
                spaceAfter=15,
            )
            section_style = ParagraphStyle(
                "Section",
                parent=styles["Heading2"],
                fontSize=13,
                textColor=colors.HexColor("#1e40af"),
                spaceBefore=12,
                spaceAfter=8,
            )
            table_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ]

            elements = [
                Paragraph(f"Ticket Performance Report - {rr.site_code}", title_style),
                Paragraph(
                    f"<b>Period:</b> {dt_from.strftime('%d %b %Y %H:%M')} to {dt_to.strftime('%d %b %Y %H:%M')}",
                    styles["Normal"],
                ),
                Spacer(1, 15),
                Paragraph("Key Metrics", section_style),
            ]

            metrics_data = [
                ["Metric", "Value"],
                ["Total Tickets", str(total_tickets)],
                [
                    "Tickets Closed",
                    f"{len(closed_tickets)} ({len(closed_tickets) / total_tickets * 100:.1f}%)"
                    if total_tickets
                    else "0 (0%)",
                ],
                ["SLA Compliance Rate", f"{sla_compliance:.1f}%"],
                ["Avg. Time to Acknowledge (MTTA)", f"{avg_mtta:.1f} Minutes"],
                ["Avg. Resolution Time (MTTR)", f"{avg_mttr:.1f} Hours"],
            ]
            elements.append(Table(metrics_data, colWidths=[200, 150]))
            elements[-1].setStyle(TableStyle(table_style))
            elements.append(Spacer(1, 15))

            elements.append(Paragraph("Tickets by Priority", section_style))
            priority_data = [["Priority", "Count"]] + [
                [p, str(c)] for p, c in priority_counts.most_common()
            ]
            elements.append(Table(priority_data, colWidths=[150, 100]))
            elements[-1].setStyle(TableStyle(table_style))
            elements.append(Spacer(1, 10))

            elements.append(Paragraph("Tickets by Status", section_style))
            status_data = [["Status", "Count"]] + [
                [s, str(c)] for s, c in status_counts.most_common()
            ]
            elements.append(Table(status_data, colWidths=[150, 100]))
            elements[-1].setStyle(TableStyle(table_style))
            elements.append(Spacer(1, 10))

            elements.append(Paragraph("Tickets by Department", section_style))
            dept_data = [["Department", "Count"]] + [
                [d, str(c)] for d, c in dept_counts.most_common(10)
            ]
            elements.append(Table(dept_data, colWidths=[200, 100]))
            elements[-1].setStyle(TableStyle(table_style))
            elements.append(Spacer(1, 10))

            elements.append(Paragraph("Top Assignees", section_style))
            user_data = [["User", "Tickets"]] + [
                [u, str(c)] for u, c in user_counts.most_common(10)
            ]
            elements.append(Table(user_data, colWidths=[200, 100]))
            elements[-1].setStyle(TableStyle(table_style))

            doc.build(elements)

        elif report_type == "sla_breach":
            # Generate Excel for SLA breach details
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"SLA_Breach_{rr.site_code}_{date_str}.xlsx"
            file_path = os.path.join(vault_root, filename)

            tickets = (
                db.execute(
                    select(Ticket).where(
                        Ticket.site_code == rr.site_code,
                        Ticket.created_at_utc >= dt_from,
                        Ticket.created_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            # Filter breached tickets
            now = _now()
            breached = []
            for t in tickets:
                if not t.sla_due_at_utc:
                    continue
                if (
                    t.status == "CLOSED"
                    and t.resolved_at_utc
                    and t.resolved_at_utc > t.sla_due_at_utc
                ):
                    breach_hours = (t.resolved_at_utc - t.sla_due_at_utc).total_seconds() / 3600
                    breached.append((t, breach_hours, "Resolved Late"))
                elif t.status != "CLOSED" and now > t.sla_due_at_utc:
                    breach_hours = (now - t.sla_due_at_utc).total_seconds() / 3600
                    breached.append((t, breach_hours, "Still Open"))

            wb = Workbook()
            ws = wb.active
            ws.title = "SLA Breaches"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

            headers = [
                "Ticket ID",
                "Asset ID",
                "Title",
                "Priority",
                "Department",
                "Created At",
                "SLA Due",
                "Resolved At",
                "Breach Hours",
                "Status",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, (t, breach_hrs, status) in enumerate(breached, 2):
                ws.cell(row=row_idx, column=1, value=t.id)
                ws.cell(row=row_idx, column=2, value=t.asset_id)
                ws.cell(row=row_idx, column=3, value=t.title[:50])
                ws.cell(row=row_idx, column=4, value=t.priority)
                ws.cell(row=row_idx, column=5, value=t.assigned_dept or "-")
                ws.cell(row=row_idx, column=6, value=t.created_at_utc.strftime("%Y-%m-%d %H:%M"))
                ws.cell(
                    row=row_idx,
                    column=7,
                    value=t.sla_due_at_utc.strftime("%Y-%m-%d %H:%M") if t.sla_due_at_utc else "-",
                )
                ws.cell(
                    row=row_idx,
                    column=8,
                    value=t.resolved_at_utc.strftime("%Y-%m-%d %H:%M")
                    if t.resolved_at_utc
                    else "-",
                )
                ws.cell(row=row_idx, column=9, value=round(breach_hrs, 1))
                ws.cell(row=row_idx, column=10, value=status)

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            wb.save(file_path)

        elif report_type == "asset_health":
            # Generate Excel for asset health metrics
            from collections import defaultdict

            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Asset_Health_{rr.site_code}_{date_str}.xlsx"
            file_path = os.path.join(vault_root, filename)

            # Get filter options
            filter_asset = filters.get("asset_id")
            critical_only = filters.get("critical_only", False)

            asset_query = select(Asset).where(
                Asset.site_code == rr.site_code, Asset.is_active.is_(True)
            )
            if filter_asset:
                asset_query = asset_query.where(Asset.id == filter_asset)
            if critical_only:
                asset_query = asset_query.where(Asset.is_critical.is_(True))
            assets = db.execute(asset_query).scalars().all()

            stops = (
                db.execute(
                    select(StopQueue).where(
                        StopQueue.site_code == rr.site_code,
                        StopQueue.opened_at_utc >= dt_from,
                        StopQueue.opened_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            tickets = (
                db.execute(
                    select(Ticket).where(
                        Ticket.site_code == rr.site_code,
                        Ticket.created_at_utc >= dt_from,
                        Ticket.created_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            # Aggregate by asset
            asset_stats = defaultdict(
                lambda: {
                    "stops": 0,
                    "downtime_min": 0,
                    "tickets": 0,
                    "name": "",
                    "category": "",
                    "is_critical": False,
                }
            )
            for a in assets:
                asset_stats[a.id]["name"] = a.name
                asset_stats[a.id]["category"] = a.category
                asset_stats[a.id]["is_critical"] = a.is_critical

            for s in stops:
                if s.asset_id in asset_stats or not filter_asset:
                    asset_stats[s.asset_id]["stops"] += 1
                    dur = ((s.closed_at_utc or dt_to) - s.opened_at_utc).total_seconds() / 60
                    asset_stats[s.asset_id]["downtime_min"] += dur

            for t in tickets:
                if t.asset_id in asset_stats or not filter_asset:
                    asset_stats[t.asset_id]["tickets"] += 1

            total_period_min = (dt_to - dt_from).total_seconds() / 60

            wb = Workbook()
            ws = wb.active
            ws.title = "Asset Health"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")

            headers = [
                "Asset ID",
                "Asset Name",
                "Category",
                "Critical",
                "Stop Count",
                "Total Downtime (Min)",
                "Avg Downtime (Min)",
                "Availability %",
                "Tickets",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            row_idx = 2
            for aid, stats in sorted(
                asset_stats.items(), key=lambda x: x[1]["downtime_min"], reverse=True
            ):
                avg_down = stats["downtime_min"] / stats["stops"] if stats["stops"] else 0
                availability = (
                    ((total_period_min - stats["downtime_min"]) / total_period_min * 100)
                    if total_period_min
                    else 100
                )
                ws.cell(row=row_idx, column=1, value=aid)
                ws.cell(row=row_idx, column=2, value=stats["name"] or aid)
                ws.cell(row=row_idx, column=3, value=stats["category"] or "-")
                ws.cell(row=row_idx, column=4, value="Yes" if stats["is_critical"] else "No")
                ws.cell(row=row_idx, column=5, value=stats["stops"])
                ws.cell(row=row_idx, column=6, value=round(stats["downtime_min"], 1))
                ws.cell(row=row_idx, column=7, value=round(avg_down, 1))
                ws.cell(row=row_idx, column=8, value=round(max(0, availability), 1))
                ws.cell(row=row_idx, column=9, value=stats["tickets"])
                row_idx += 1

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

            wb.save(file_path)

        elif report_type == "stop_reason_analysis":
            # Generate PDF for stop reason Pareto analysis
            from collections import Counter

            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Stop_Reason_Analysis_{rr.site_code}_{date_str}.pdf"
            file_path = os.path.join(vault_root, filename)

            stops = (
                db.execute(
                    select(StopQueue).where(
                        StopQueue.site_code == rr.site_code,
                        StopQueue.opened_at_utc >= dt_from,
                        StopQueue.opened_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            reason_freq = Counter(s.reason or "Unknown" for s in stops)
            reason_downtime = {}
            for s in stops:
                r = s.reason or "Unknown"
                dur = ((s.closed_at_utc or dt_to) - s.opened_at_utc).total_seconds() / 60
                reason_downtime[r] = reason_downtime.get(r, 0) + dur

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "Title",
                parent=styles["Heading1"],
                fontSize=20,
                textColor=colors.HexColor("#1e40af"),
                spaceAfter=15,
            )
            section_style = ParagraphStyle(
                "Section",
                parent=styles["Heading2"],
                fontSize=13,
                textColor=colors.HexColor("#1e40af"),
                spaceBefore=12,
                spaceAfter=8,
            )
            table_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ]

            elements = [
                Paragraph(f"Stop Reason Analysis - {rr.site_code}", title_style),
                Paragraph(
                    f"<b>Period:</b> {dt_from.strftime('%d %b %Y')} to {dt_to.strftime('%d %b %Y')}",
                    styles["Normal"],
                ),
                Paragraph(f"<b>Total Stops:</b> {len(stops)}", styles["Normal"]),
                Spacer(1, 15),
            ]

            # By Frequency
            elements.append(Paragraph("Top Reasons by Frequency", section_style))
            total_stops = len(stops)
            freq_data = [["Rank", "Reason", "Count", "% of Total"]]
            cumulative = 0
            for i, (reason, cnt) in enumerate(reason_freq.most_common(15), 1):
                pct = cnt / total_stops * 100 if total_stops else 0
                cumulative += pct
                freq_data.append([str(i), reason[:40], str(cnt), f"{pct:.1f}%"])
            elements.append(Table(freq_data, colWidths=[40, 220, 60, 80]))
            elements[-1].setStyle(TableStyle(table_style))
            elements.append(Spacer(1, 15))

            # By Downtime
            elements.append(Paragraph("Top Reasons by Downtime", section_style))
            sorted_by_down = sorted(reason_downtime.items(), key=lambda x: x[1], reverse=True)[:15]
            total_down = sum(reason_downtime.values())
            down_data = [["Rank", "Reason", "Downtime (Min)", "% of Total"]]
            for i, (reason, mins) in enumerate(sorted_by_down, 1):
                pct = mins / total_down * 100 if total_down else 0
                down_data.append([str(i), reason[:40], f"{mins:.0f}", f"{pct:.1f}%"])
            elements.append(Table(down_data, colWidths=[40, 220, 100, 80]))
            elements[-1].setStyle(TableStyle(table_style))

            doc.build(elements)

        elif report_type == "personnel_performance":
            # Generate Excel for maintenance personnel metrics
            from collections import defaultdict

            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Personnel_Performance_{rr.site_code}_{date_str}.xlsx"
            file_path = os.path.join(vault_root, filename)

            filter_user = filters.get("user_id")

            tickets = (
                db.execute(
                    select(Ticket).where(
                        Ticket.site_code == rr.site_code,
                        Ticket.created_at_utc >= dt_from,
                        Ticket.created_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            if filter_user:
                tickets = [t for t in tickets if t.assigned_to_user_id == filter_user]

            user_stats = defaultdict(lambda: {"assigned": 0, "closed": 0, "resolution_times": []})
            for t in tickets:
                user = t.assigned_to_user_id or "Unassigned"
                user_stats[user]["assigned"] += 1
                if t.status == "CLOSED":
                    user_stats[user]["closed"] += 1
                    if t.resolved_at_utc and t.created_at_utc:
                        user_stats[user]["resolution_times"].append(
                            (t.resolved_at_utc - t.created_at_utc).total_seconds() / 3600
                        )

            wb = Workbook()
            ws = wb.active
            ws.title = "Personnel Performance"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

            headers = [
                "User ID",
                "Tickets Assigned",
                "Tickets Closed",
                "Resolution Rate %",
                "Avg Resolution Time (Hours)",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            row_idx = 2
            for user, stats in sorted(
                user_stats.items(), key=lambda x: x[1]["assigned"], reverse=True
            ):
                res_rate = stats["closed"] / stats["assigned"] * 100 if stats["assigned"] else 0
                avg_res = (
                    sum(stats["resolution_times"]) / len(stats["resolution_times"])
                    if stats["resolution_times"]
                    else 0
                )
                ws.cell(row=row_idx, column=1, value=user)
                ws.cell(row=row_idx, column=2, value=stats["assigned"])
                ws.cell(row=row_idx, column=3, value=stats["closed"])
                ws.cell(row=row_idx, column=4, value=round(res_rate, 1))
                ws.cell(row=row_idx, column=5, value=round(avg_res, 1))
                row_idx += 1

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

            wb.save(file_path)

        elif report_type == "critical_asset":
            # Generate PDF for critical asset focus
            from collections import defaultdict

            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Critical_Asset_{rr.site_code}_{date_str}.pdf"
            file_path = os.path.join(vault_root, filename)

            critical_assets = (
                db.execute(
                    select(Asset).where(
                        Asset.site_code == rr.site_code,
                        Asset.is_active.is_(True),
                        Asset.is_critical.is_(True),
                    )
                )
                .scalars()
                .all()
            )

            stops = (
                db.execute(
                    select(StopQueue).where(
                        StopQueue.site_code == rr.site_code,
                        StopQueue.opened_at_utc >= dt_from,
                        StopQueue.opened_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            tickets = (
                db.execute(
                    select(Ticket).where(
                        Ticket.site_code == rr.site_code,
                        Ticket.created_at_utc >= dt_from,
                        Ticket.created_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            critical_ids = {a.id for a in critical_assets}
            critical_stops = [s for s in stops if s.asset_id in critical_ids]
            critical_tickets = [t for t in tickets if t.asset_id in critical_ids]

            asset_stats = defaultdict(
                lambda: {"name": "", "stops": 0, "downtime": 0, "tickets": 0, "open_tickets": 0}
            )
            for a in critical_assets:
                asset_stats[a.id]["name"] = a.name

            for s in critical_stops:
                asset_stats[s.asset_id]["stops"] += 1
                dur = ((s.closed_at_utc or dt_to) - s.opened_at_utc).total_seconds() / 60
                asset_stats[s.asset_id]["downtime"] += dur

            for t in critical_tickets:
                asset_stats[t.asset_id]["tickets"] += 1
                if t.status != "CLOSED":
                    asset_stats[t.asset_id]["open_tickets"] += 1

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "Title",
                parent=styles["Heading1"],
                fontSize=20,
                textColor=colors.HexColor("#DC2626"),
                spaceAfter=15,
            )
            section_style = ParagraphStyle(
                "Section",
                parent=styles["Heading2"],
                fontSize=13,
                textColor=colors.HexColor("#DC2626"),
                spaceBefore=12,
                spaceAfter=8,
            )
            table_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DC2626")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#fecaca")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fef2f2")]),
            ]

            elements = [
                Paragraph(f"Critical Asset Report - {rr.site_code}", title_style),
                Paragraph(
                    f"<b>Period:</b> {dt_from.strftime('%d %b %Y')} to {dt_to.strftime('%d %b %Y')}",
                    styles["Normal"],
                ),
                Spacer(1, 15),
                Paragraph("Summary", section_style),
            ]

            summary_data = [
                ["Metric", "Value"],
                ["Critical Assets Monitored", str(len(critical_assets))],
                ["Total Stops on Critical", str(len(critical_stops))],
                [
                    "Total Downtime (Hours)",
                    f"{sum(s['downtime'] for s in asset_stats.values()) / 60:.1f}",
                ],
                [
                    "Open Tickets on Critical",
                    str(sum(s["open_tickets"] for s in asset_stats.values())),
                ],
            ]
            elements.append(Table(summary_data, colWidths=[200, 150]))
            elements[-1].setStyle(TableStyle(table_style))
            elements.append(Spacer(1, 15))

            elements.append(Paragraph("Critical Asset Details", section_style))
            detail_data = [["Asset ID", "Name", "Stops", "Downtime (Min)", "Tickets", "Open"]]
            for aid, s in sorted(asset_stats.items(), key=lambda x: x[1]["downtime"], reverse=True):
                detail_data.append(
                    [
                        aid,
                        s["name"][:25],
                        str(s["stops"]),
                        f"{s['downtime']:.0f}",
                        str(s["tickets"]),
                        str(s["open_tickets"]),
                    ]
                )

            if len(detail_data) == 1:
                detail_data.append(["No critical assets", "-", "-", "-", "-", "-"])

            elements.append(Table(detail_data, colWidths=[80, 130, 50, 80, 50, 40]))
            elements[-1].setStyle(TableStyle(table_style))

            doc.build(elements)

        elif report_type == "department_performance":
            # Generate Excel for department metrics
            from collections import defaultdict

            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Department_Performance_{rr.site_code}_{date_str}.xlsx"
            file_path = os.path.join(vault_root, filename)

            tickets = (
                db.execute(
                    select(Ticket).where(
                        Ticket.site_code == rr.site_code,
                        Ticket.created_at_utc >= dt_from,
                        Ticket.created_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            dept_stats = defaultdict(
                lambda: {"total": 0, "closed": 0, "sla_met": 0, "resolution_times": []}
            )
            for t in tickets:
                dept = t.assigned_dept or "Unassigned"
                dept_stats[dept]["total"] += 1
                if t.status == "CLOSED":
                    dept_stats[dept]["closed"] += 1
                    if (
                        t.sla_due_at_utc
                        and t.resolved_at_utc
                        and t.resolved_at_utc <= t.sla_due_at_utc
                    ):
                        dept_stats[dept]["sla_met"] += 1
                    if t.resolved_at_utc and t.created_at_utc:
                        dept_stats[dept]["resolution_times"].append(
                            (t.resolved_at_utc - t.created_at_utc).total_seconds() / 3600
                        )

            wb = Workbook()
            ws = wb.active
            ws.title = "Department Performance"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")

            headers = [
                "Department",
                "Total Tickets",
                "Closed",
                "Resolution Rate %",
                "SLA Compliance %",
                "Avg MTTR (Hours)",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            row_idx = 2
            for dept, stats in sorted(
                dept_stats.items(), key=lambda x: x[1]["total"], reverse=True
            ):
                res_rate = stats["closed"] / stats["total"] * 100 if stats["total"] else 0
                sla_rate = stats["sla_met"] / stats["closed"] * 100 if stats["closed"] else 0
                avg_mttr = (
                    sum(stats["resolution_times"]) / len(stats["resolution_times"])
                    if stats["resolution_times"]
                    else 0
                )
                ws.cell(row=row_idx, column=1, value=dept)
                ws.cell(row=row_idx, column=2, value=stats["total"])
                ws.cell(row=row_idx, column=3, value=stats["closed"])
                ws.cell(row=row_idx, column=4, value=round(res_rate, 1))
                ws.cell(row=row_idx, column=5, value=round(sla_rate, 1))
                ws.cell(row=row_idx, column=6, value=round(avg_mttr, 1))
                row_idx += 1

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

            wb.save(file_path)

        elif report_type == "audit_trail":
            # Generate Excel for audit log export
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Audit_Trail_{rr.site_code}_{date_str}.xlsx"
            file_path = os.path.join(vault_root, filename)

            filter_entity = filters.get("entity_type")
            filter_user = filters.get("user_id")

            audit_query = select(AuditLog).where(
                AuditLog.site_code == rr.site_code,
                AuditLog.created_at_utc >= dt_from,
                AuditLog.created_at_utc <= dt_to,
            )
            if filter_entity:
                audit_query = audit_query.where(AuditLog.entity_type == filter_entity)
            if filter_user:
                audit_query = audit_query.where(AuditLog.actor_user_id == filter_user)

            logs = db.execute(audit_query.order_by(AuditLog.created_at_utc.desc())).scalars().all()

            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Trail"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")

            headers = ["Timestamp", "User", "Action", "Entity Type", "Entity ID", "Details"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            import json

            for row_idx, log in enumerate(logs, 2):
                ws.cell(
                    row=row_idx, column=1, value=log.created_at_utc.strftime("%Y-%m-%d %H:%M:%S")
                )
                ws.cell(row=row_idx, column=2, value=log.actor_user_id or "SYSTEM")
                ws.cell(row=row_idx, column=3, value=log.action)
                ws.cell(row=row_idx, column=4, value=log.entity_type)
                ws.cell(row=row_idx, column=5, value=log.entity_id)
                details = (
                    log.details_json
                    if isinstance(log.details_json, str)
                    else json.dumps(log.details_json)
                )
                ws.cell(row=row_idx, column=6, value=details[:200])

            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            wb.save(file_path)

        elif report_type == "trend_analysis":
            # Generate PDF for trend analysis
            from collections import defaultdict

            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Trend_Analysis_{rr.site_code}_{date_str}.pdf"
            file_path = os.path.join(vault_root, filename)

            stops = (
                db.execute(
                    select(StopQueue).where(
                        StopQueue.site_code == rr.site_code,
                        StopQueue.opened_at_utc >= dt_from,
                        StopQueue.opened_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            tickets = (
                db.execute(
                    select(Ticket).where(
                        Ticket.site_code == rr.site_code,
                        Ticket.created_at_utc >= dt_from,
                        Ticket.created_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            # Daily aggregation
            daily_stops = defaultdict(int)
            daily_downtime = defaultdict(float)
            daily_tickets = defaultdict(int)
            daily_closed = defaultdict(int)

            for s in stops:
                day = s.opened_at_utc.strftime("%Y-%m-%d")
                daily_stops[day] += 1
                dur = ((s.closed_at_utc or dt_to) - s.opened_at_utc).total_seconds() / 60
                daily_downtime[day] += dur

            for t in tickets:
                day = t.created_at_utc.strftime("%Y-%m-%d")
                daily_tickets[day] += 1
                if t.status == "CLOSED":
                    daily_closed[day] += 1

            all_days = sorted(set(daily_stops.keys()) | set(daily_tickets.keys()))

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "Title",
                parent=styles["Heading1"],
                fontSize=20,
                textColor=colors.HexColor("#1e40af"),
                spaceAfter=15,
            )
            section_style = ParagraphStyle(
                "Section",
                parent=styles["Heading2"],
                fontSize=13,
                textColor=colors.HexColor("#1e40af"),
                spaceBefore=12,
                spaceAfter=8,
            )
            table_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ]

            elements = [
                Paragraph(f"Trend Analysis Report - {rr.site_code}", title_style),
                Paragraph(
                    f"<b>Period:</b> {dt_from.strftime('%d %b %Y')} to {dt_to.strftime('%d %b %Y')}",
                    styles["Normal"],
                ),
                Paragraph(f"<b>Days Analyzed:</b> {len(all_days)}", styles["Normal"]),
                Spacer(1, 15),
                Paragraph("Overall Trends", section_style),
            ]

            total_stops = sum(daily_stops.values())
            total_tickets = sum(daily_tickets.values())
            avg_daily_stops = total_stops / len(all_days) if all_days else 0
            avg_daily_tickets = total_tickets / len(all_days) if all_days else 0

            summary_data = [
                ["Metric", "Value"],
                ["Total Days", str(len(all_days))],
                ["Total Stops", str(total_stops)],
                ["Avg Stops/Day", f"{avg_daily_stops:.1f}"],
                ["Total Tickets", str(total_tickets)],
                ["Avg Tickets/Day", f"{avg_daily_tickets:.1f}"],
            ]
            elements.append(Table(summary_data, colWidths=[180, 120]))
            elements[-1].setStyle(TableStyle(table_style))
            elements.append(Spacer(1, 15))

            elements.append(Paragraph("Daily Breakdown", section_style))
            trend_data = [["Date", "Stops", "Downtime (Min)", "Tickets", "Closed"]]
            for day in all_days[-30:]:  # Last 30 days max
                trend_data.append(
                    [
                        day,
                        str(daily_stops.get(day, 0)),
                        f"{daily_downtime.get(day, 0):.0f}",
                        str(daily_tickets.get(day, 0)),
                        str(daily_closed.get(day, 0)),
                    ]
                )

            if len(trend_data) == 1:
                trend_data.append(["No data", "0", "0", "0", "0"])

            elements.append(Table(trend_data, colWidths=[80, 50, 80, 50, 50]))
            elements[-1].setStyle(TableStyle(table_style))

            doc.build(elements)

        else:
            # Generate PDF for daily_summary (default)
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            date_str = f"{dt_from.strftime('%d%b%y')}_to_{dt_to.strftime('%d%b%y')}"
            filename = f"Summary_{rr.site_code}_{date_str}.pdf"
            file_path = os.path.join(vault_root, filename)

            stops = (
                db.execute(
                    select(StopQueue).where(
                        StopQueue.site_code == rr.site_code,
                        StopQueue.opened_at_utc >= dt_from,
                        StopQueue.opened_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            tickets = (
                db.execute(
                    select(Ticket).where(
                        Ticket.site_code == rr.site_code,
                        Ticket.created_at_utc >= dt_from,
                        Ticket.created_at_utc <= dt_to,
                    )
                )
                .scalars()
                .all()
            )

            # Calculate basic stats
            total_stops = len(stops)
            total_downtime_min = sum(
                (((s.closed_at_utc or dt_to) - s.opened_at_utc).total_seconds() / 60) for s in stops
            )
            total_tickets = len(tickets)
            closed_tickets = len([t for t in tickets if t.status == "CLOSED"])

            # Calculate advanced stats
            from collections import Counter, defaultdict

            asset_downtime = defaultdict(float)
            for s in stops:
                dur = ((s.closed_at_utc or dt_to) - s.opened_at_utc).total_seconds() / 60
                asset_downtime[s.asset_id] += dur
            top_assets = sorted(asset_downtime.items(), key=lambda x: x[1], reverse=True)[:10]

            reason_counts = Counter(s.reason or "Unknown" for s in stops)
            top_reasons = reason_counts.most_common(10)

            priority_counts = Counter(t.priority for t in tickets)
            status_counts = Counter(t.status for t in tickets)

            # MTTR calculation
            resolved_diffs = []
            for t in tickets:
                if t.status == "CLOSED" and t.resolved_at_utc and t.created_at_utc:
                    resolved_diffs.append(
                        (t.resolved_at_utc - t.created_at_utc).total_seconds() / 3600
                    )
            avg_mttr = sum(resolved_diffs) / len(resolved_diffs) if resolved_diffs else 0

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "Title",
                parent=styles["Heading1"],
                fontSize=22,
                textColor=colors.HexColor("#1e40af"),
                spaceAfter=20,
            )
            section_style = ParagraphStyle(
                "Section",
                parent=styles["Heading2"],
                fontSize=14,
                textColor=colors.HexColor("#1e40af"),
                spaceBefore=15,
                spaceAfter=10,
            )
            table_header_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f3f4f6")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ]

            elements = []

            # 1. Title & Header
            elements.append(Paragraph(f"Management Summary Report - {rr.site_code}", title_style))
            elements.append(
                Paragraph(
                    f"<b>Reporting Period:</b> {dt_from.strftime('%d %b %Y %H:%M')} to {dt_to.strftime('%d %b %Y %H:%M')}",
                    styles["Normal"],
                )
            )
            elements.append(
                Paragraph(
                    f"<b>Report Generated:</b> {_now().strftime('%d %b %Y %H:%M:%S')} UTC",
                    styles["Normal"],
                )
            )
            elements.append(Spacer(1, 20))

            # 2. Executive Summary Metrics
            elements.append(Paragraph("Executive Overview", section_style))
            summary_data = [
                ["KPI Metric", "Value"],
                ["Total Downtime Events", str(total_stops)],
                ["Total Cumulative Downtime", f"{total_downtime_min:.1f} Minutes"],
                ["Total Support Tickets", str(total_tickets)],
                [
                    "Tickets Resolved",
                    f"{closed_tickets} ({(closed_tickets / total_tickets * 100 if total_tickets else 0):.1f}%)",
                ],
                ["Avg. Resolution Time (MTTR)", f"{avg_mttr:.1f} Hours"],
            ]
            t_summary = Table(summary_data, colWidths=[250, 150])
            t_summary.setStyle(TableStyle(table_header_style))
            elements.append(t_summary)
            elements.append(Spacer(1, 20))

            # 3. Top 10 Assets by Downtime
            elements.append(Paragraph("Top 10 Assets by Cumulative Downtime", section_style))
            asset_data = [["Asset ID", "Downtime (Min)"]]
            for asset_id, mins in top_assets:
                asset_data.append([asset_id, f"{mins:.1f}"])
            if len(top_assets) == 0:
                asset_data.append(["No data", "0.0"])

            t_assets = Table(asset_data, colWidths=[250, 150])
            t_assets.setStyle(TableStyle(table_header_style))
            elements.append(t_assets)
            elements.append(Spacer(1, 20))

            # 4. Stop Reason Analysis
            elements.append(Paragraph("Top Stop Reasons", section_style))
            reason_data = [["Stop Reason", "Frequency"]]
            for reason, count in top_reasons:
                reason_data.append([reason, str(count)])
            if len(top_reasons) == 0:
                reason_data.append(["No data", "0"])

            t_reasons = Table(reason_data, colWidths=[250, 150])
            t_reasons.setStyle(TableStyle(table_header_style))
            elements.append(t_reasons)

            # 5. Ticket Distribution (Side-by-Side Table)
            elements.append(Paragraph("Ticket Distribution", section_style))
            ticket_data = [
                ["By Priority", "Count", "", "By Status", "Count"],
                ["High", str(priority_counts["HIGH"]), "", "Open", str(status_counts["OPEN"])],
                [
                    "Medium",
                    str(priority_counts["MEDIUM"]),
                    "",
                    "Acknowledged",
                    str(status_counts["ACKNOWLEDGED"]),
                ],
                ["Low", str(priority_counts["LOW"]), "", "Resolved", str(status_counts["CLOSED"])],
            ]
            t_tickets = Table(ticket_data, colWidths=[100, 50, 20, 100, 50])
            t_tickets.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#2563EB")),
                        ("BACKGROUND", (3, 0), (4, 0), colors.HexColor("#2563EB")),
                        ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                        ("TEXTCOLOR", (3, 0), (4, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("GRID", (0, 0), (1, -1), 0.5, colors.HexColor("#e5e7eb")),
                        ("GRID", (3, 0), (4, -1), 0.5, colors.HexColor("#e5e7eb")),
                    ]
                )
            )
            elements.append(t_tickets)

            doc.build(elements)

        rr.status = "generated"
        rr.generated_file_path = filename
        rr.updated_at_utc = _now()
        audit_write(
            db,
            "REPORT_GENERATED",
            "report_request",
            str(rr.id),
            {"path": filename},
            actor_user_id,
            actor_station_code,
            request_id,
        )
    except Exception as e:
        rr.status = "failed"
        rr.error_message = str(e)
        rr.updated_at_utc = _now()
        audit_write(
            db,
            "REPORT_FAILED",
            "report_request",
            str(rr.id),
            {"error": str(e)},
            actor_user_id,
            actor_station_code,
            request_id,
        )

    return rr


def check_sla_warnings(db) -> int:
    """
    Check for tickets approaching SLA deadline and queue warning alerts.
    Call this periodically from the plant_worker.
    Returns the number of warnings sent.
    """
    import logging

    log = logging.getLogger("assetiq")

    now = _now()
    warning_threshold = now + timedelta(hours=1)  # Warning if SLA due within 1 hour

    # Find open tickets approaching SLA that haven't had warning sent
    from sqlalchemy import and_

    tickets = (
        db.execute(
            select(Ticket).where(
                and_(
                    Ticket.status.in_(["OPEN", "ACKNOWLEDGED", "ACK"]),
                    Ticket.sla_due_at_utc.isnot(None),
                    Ticket.sla_due_at_utc <= warning_threshold,
                    Ticket.sla_due_at_utc > now,  # Not yet breached
                    Ticket.sla_warning_sent.is_(False),
                )
            )
        )
        .scalars()
        .all()
    )

    if not tickets:
        return 0

    # Check if WhatsApp is enabled
    ws_enabled = db.get(SystemConfig, "whatsappEnabled")
    ws_phone = db.get(SystemConfig, "whatsappTargetPhone")

    if not (ws_enabled and ws_enabled.config_value is True and ws_phone and ws_phone.config_value):
        return 0

    count = 0
    for t in tickets:
        try:
            # Calculate remaining time
            remaining = t.sla_due_at_utc - now
            remaining_mins = int(remaining.total_seconds() / 60)

            # Build warning message
            msg = (
                f"⚠️ SLA Warning\n"
                f"Ticket: {t.id}\n"
                f"Asset: {t.asset_id}\n"
                f"Title: {t.title}\n"
                f"Priority: {t.priority}\n"
                f"Time Remaining: {remaining_mins} minutes\n"
                f"SLA Due: {t.sla_due_at_utc.strftime('%H:%M')}"
            )

            db.add(
                WhatsAppQueue(
                    ticket_id=t.id,
                    phone_number=str(ws_phone.config_value),
                    message=msg,
                    status="PENDING",
                    sla_state="WARNING",
                    created_at_utc=now,
                )
            )

            # Mark warning as sent
            t.sla_warning_sent = True
            count += 1

        except Exception as e:
            log.error(f"Failed to queue SLA warning for ticket {t.id}: {e}")

    db.commit()
    log.info(f"sla_warnings_queued count={count}")
    return count
